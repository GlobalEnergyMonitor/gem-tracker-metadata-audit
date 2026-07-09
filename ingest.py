"""
ingest.py

Loads data tabs from all .xlsx files in gem-data/ into tracker_data.db (SQLite).
All values are stored as TEXT; type inference happens at analysis time.

Also loads extracted metadata JSON (from metadata/) into the _tracker_fields table
if available. Run extract_metadata.py first to generate those JSON files.

Run: python3 ingest.py
"""

import itertools
import json
import re
import sqlite3
from pathlib import Path

import openpyxl

SPREADSHEET_DIR = Path("gem-data")
METADATA_DIR = Path("metadata")
DB_PATH = "tracker_data.db"

METADATA_TAB_NAMES = {
    "readme", "about", "metadata", "notes", "dictionary", "instructions",
    "data dictionary", "acronyms", "copyright", "introduction terminology",
    "column key",
}


def slugify(text):
    text = str(text).lower().strip()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_")


def normalize_col(name):
    """Strip newlines and collapse whitespace in a column name."""
    if name is None:
        return None
    return re.sub(r"\s+", " ", str(name).strip())


def create_schema(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS _tracker_fields (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            tracker_slug  TEXT NOT NULL,
            tab_slug      TEXT NOT NULL,
            section_name  TEXT,
            field_name    TEXT NOT NULL,
            definition    TEXT,
            notes         TEXT,
            entry_type    TEXT,
            parent_field  TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_tf_tracker ON _tracker_fields(tracker_slug);
        CREATE INDEX IF NOT EXISTS idx_tf_field   ON _tracker_fields(tracker_slug, field_name);

        CREATE TABLE IF NOT EXISTS _trackers (
            tracker_slug  TEXT PRIMARY KEY,
            source_file   TEXT,
            title         TEXT,
            citation      TEXT,
            contact       TEXT,
            license       TEXT
        );
    """)
    conn.commit()


def load_metadata_json(conn, tracker_slug, metadata_path):
    """Populate _tracker_fields and _trackers from an extracted metadata JSON file."""
    if not metadata_path.exists():
        return

    with open(metadata_path, encoding="utf-8") as f:
        data = json.load(f)

    hi = data.get("header_info", {})
    conn.execute(
        """
        INSERT OR REPLACE INTO _trackers (tracker_slug, source_file, title, citation, contact, license)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            tracker_slug,
            data.get("source_file"),
            hi.get("title"),
            hi.get("citation"),
            hi.get("contact"),
            hi.get("license"),
        ),
    )

    conn.execute("DELETE FROM _tracker_fields WHERE tracker_slug = ?", (tracker_slug,))

    rows = []
    for section in data.get("sections", []):
        for entry in section.get("entries", []):
            rows.append(
                (
                    tracker_slug,
                    section.get("tab_key") or "unknown",
                    section.get("section_name"),
                    entry["name"],
                    entry.get("definition"),
                    entry.get("notes"),
                    entry.get("entry_type"),
                    entry.get("parent_field"),
                )
            )

    conn.executemany(
        """
        INSERT INTO _tracker_fields
            (tracker_slug, tab_slug, section_name, field_name, definition, notes, entry_type, parent_field)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    conn.commit()
    print(f"  Metadata: {len(rows)} entries loaded into _tracker_fields")


def load_data_tab(conn, ws, table_name):
    """
    Stream a worksheet into a SQLite table (all columns TEXT).
    Returns the number of data rows inserted.
    """
    rows_iter = ws.iter_rows(values_only=True)

    # Header row — skip a leading title row if only one cell is non-empty
    try:
        raw_header = next(rows_iter)
    except StopIteration:
        return 0

    non_null_count = sum(1 for c in raw_header if c is not None and str(c).strip())
    n_cols = len(raw_header)
    # Skip a leading section-label row: very sparse (<10% filled, <3 cells) relative to column count
    if n_cols > 2 and non_null_count < max(3, n_cols * 0.1):
        try:
            next_row = next(rows_iter)
        except StopIteration:
            pass
        else:
            non_null_next = sum(1 for c in next_row if c is not None and str(c).strip())
            if non_null_next > non_null_count:
                raw_header = next_row
            else:
                rows_iter = itertools.chain([next_row], rows_iter)

    raw_names = [normalize_col(c) or f"_col{i}" for i, c in enumerate(raw_header)]

    # Deduplicate column names by appending _2, _3, etc.
    seen: dict[str, int] = {}
    header = []
    for name in raw_names:
        if name not in seen:
            seen[name] = 1
            header.append(name)
        else:
            seen[name] += 1
            header.append(f"{name}_{seen[name]}")

    n_cols = len(header)
    col_defs = ", ".join(f'"{h}" TEXT' for h in header)
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
    conn.commit()

    placeholders = ", ".join("?" * n_cols)
    insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'

    batch = []
    n_rows = 0

    for row in rows_iter:
        vals = [str(c) if c is not None else None for c in row[:n_cols]]
        # Pad if row is shorter than header (sparse trailing cells)
        while len(vals) < n_cols:
            vals.append(None)
        batch.append(vals)
        n_rows += 1

        if len(batch) >= 500:
            conn.executemany(insert_sql, batch)
            conn.commit()
            batch = []

    if batch:
        conn.executemany(insert_sql, batch)
        conn.commit()

    return n_rows


def process_file(conn, filepath):
    tracker_slug = slugify(filepath.stem)
    print(f"\nIngesting: {filepath.name}  (slug: {tracker_slug})")

    # Use read_only for memory efficiency on large files
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)

    for tab_name in wb.sheetnames:
        if tab_name.lower() in METADATA_TAB_NAMES:
            continue

        tab_slug = slugify(tab_name)
        table_name = f"{tracker_slug}__{tab_slug}"
        ws = wb[tab_name]

        n = load_data_tab(conn, ws, table_name)
        print(f"  Table '{table_name}': {n:,} rows")

    wb.close()

    # Load metadata if extract_metadata.py has been run
    metadata_path = METADATA_DIR / f"{tracker_slug}.json"
    load_metadata_json(conn, tracker_slug, metadata_path)


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    create_schema(conn)

    xlsx_files = sorted(SPREADSHEET_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files found in {SPREADSHEET_DIR}/")
        conn.close()
        return

    for filepath in xlsx_files:
        process_file(conn, filepath)

    # Summary
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE '\_%' ESCAPE '\\' ORDER BY name"
    ).fetchall()
    print(f"\nDone. {DB_PATH} has {len(tables)} data table(s).")
    conn.close()


if __name__ == "__main__":
    main()
