"""
extract_metadata.py

Reads metadata tab(s) from each .xlsx in gem-data/, parses field definitions
and category values, and writes structured JSON to metadata/{tracker_slug}.json.

Section types handled:
  'dictionary'         — coal mine/plant README format (col_a=field, col_b=def)
  'columns'            — standard power tracker "Columns" section (col_b=units, col_c=def)
  'definitions'        — "Definitions" / "Status Definitions" sections (value definitions)
  'flat'               — dedicated Metadata/Data Dictionary tabs or auto-detected
  'column_key_tabbed'  — Oil Extraction "Column Key" style (col_a=tab, col_b=field, col_c=def)

Handles layouts where all content is shifted one column right due to merged cells.

Run: python3 extract_metadata.py
"""

import json
import re
from pathlib import Path
import openpyxl

SPREADSHEET_DIR = Path("gem-data")
OUTPUT_DIR = Path("metadata")

# Tab names treated as metadata (not data) — keep in sync with ingest.py
METADATA_TAB_NAMES = {
    "readme", "about", "metadata", "notes", "dictionary", "instructions",
    "data dictionary", "acronyms", "copyright", "introduction terminology",
    "column key",
}

# Tabs from which we extract header info (title, citation, contact, license)
HEADER_INFO_TAB_NAMES = {"readme", "about", "notes"}

# Section type keywords
SECTION_KEYWORDS = ("dictionary", "field description", "field descriptions")
COLUMNS_SECTION_KEYWORDS = ("columns",)
DEFINITIONS_SECTION_KEYWORDS = (
    "definitions", "status definitions", "attribute explanations",
)

# First-cell values that mark a structural sub-header row to skip
COLUMN_HEADER_FIRST_CELLS = {
    "column name", "column", "field name", "field", "name",
    "variable name", "variablename", "variable",
    "metadata",   # "Metadata | Description" header in Cement
    "sheet",      # "Sheet | Column name | Definition" in Oil Extraction
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def slugify(text):
    text = str(text).lower().strip()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_")


def normalize(text):
    return re.sub(r"\s+", " ", str(text).lower().strip())


def cell_str(row, idx):
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def normalize_row(row):
    """
    Some tabs shift all content one column right due to merged cells.
    If col A is empty and col B has content, shift the whole row left by one.
    """
    if not row:
        return row
    col_a = row[0]
    col_b = row[1] if len(row) > 1 else None
    if (col_a is None or str(col_a).strip() == "") and col_b is not None and str(col_b).strip():
        return row[1:]
    return row


def extract_header_info(rows):
    """Pull attribution fields from the first ~20 rows."""
    info = {"title": None, "citation": None, "contact": None, "license": None}
    for raw_row in rows[:20]:
        row = normalize_row(raw_row)
        # Try col A, fall back to col B
        text = None
        for idx in (0, 1):
            val = row[idx] if len(row) > idx else None
            if val is not None:
                s = str(val).strip().replace("\n", " ")
                if s:
                    text = s
                    break
        if not text:
            continue
        lower = text.lower()
        if "copyright" in lower and not info["license"]:
            info["license"] = text
        elif ("citation" in lower or "recommended" in lower) and not info["citation"]:
            info["citation"] = text
        elif "contact" in lower and not info["contact"]:
            info["contact"] = text
        elif not info["title"] and len(text) > 8:
            info["title"] = text
    return info


def is_column_header_row(row):
    col_a = cell_str(row, 0).lower()
    return col_a in COLUMN_HEADER_FIRST_CELLS


def infer_tab_key(hint_text, tab_keys):
    """Match a freetext hint to a known tab slug."""
    def match(slug):
        if slug in tab_keys:
            return slug
        for key in tab_keys:
            if key.startswith(slug) or slug.startswith(key):
                return key
        return None

    # "(Non-closed mines)" style
    m = re.search(r"\((.+?)\)", hint_text)
    if m:
        found = match(slugify(m.group(1)))
        if found:
            return found
    # "Columns for Data tab" style
    m = re.search(r"\bfor\s+(.+?)(?:\s+tab)?$", hint_text, re.IGNORECASE)
    if m:
        found = match(slugify(m.group(1)))
        if found:
            return found

    return tab_keys[0] if len(tab_keys) == 1 else None


def get_section_info(row, tab_keys):
    """
    If this row is a section header, return (section_name, section_type, tab_key).
    Otherwise return None.
    """
    col_a = cell_str(row, 0)
    col_b = cell_str(row, 1)
    if not col_a or col_b:
        return None  # section headers have col_a content, col_b empty

    lower = col_a.lower()

    # "Column Key" / Oil Extraction multi-tab style
    if lower == "column key" or lower.startswith("column key "):
        return col_a, "column_key_tabbed", None

    if any(kw in lower for kw in SECTION_KEYWORDS):
        return col_a, "dictionary", infer_tab_key(col_a, tab_keys)

    if any(lower.startswith(kw) for kw in COLUMNS_SECTION_KEYWORDS):
        return col_a, "columns", infer_tab_key(col_a, tab_keys)

    if any(kw in lower for kw in DEFINITIONS_SECTION_KEYWORDS):
        return col_a, "definitions", None

    # '"Plant data" tab' style (Iron/Steel Metadata tab)
    m = re.match(r'^"([^"]+)"\s+tab$', col_a, re.IGNORECASE)
    if m:
        tab_key = infer_tab_key(slugify(m.group(1)), tab_keys)
        return col_a, "flat", tab_key

    # Bare tab slug (GMET Data Dictionary: "Plumes", "Pipelines", etc.)
    slug = slugify(col_a)
    if slug in tab_keys:
        return col_a, "flat", slug

    return None


def infer_parent_field(section_name):
    """'Status Definitions' → 'Status'; 'Definitions' → None."""
    lower = section_name.lower().strip()
    for suffix in (" definitions", " definition", " explanations", " explanation"):
        if lower.endswith(suffix):
            parent = section_name[: -len(suffix)].strip()
            return parent if parent else None
    return None


def close_section(sections, current_section):
    """Finalize and append a section, stripping internal tracking keys."""
    if current_section is None:
        return
    current_section.pop("_last_tab", None)
    if current_section["entries"]:
        sections.append(current_section)


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_metadata_tab(ws, tab_col_sets):
    """
    Parse one metadata worksheet.

    Returns (header_info, sections) where sections is a list of:
      {section_name, section_type, tab_key, entries}
    and each entry is:
      {name, definition, notes, entry_type, parent_field}
    """
    rows = list(ws.iter_rows(values_only=True))
    header_info = extract_header_info(rows)

    all_col_norms = {}
    for tab_slug, col_set in tab_col_sets.items():
        for norm in col_set:
            all_col_norms[norm] = tab_slug

    tab_keys = list(tab_col_sets.keys())
    sections = []
    current_section = None
    current_section_type = None
    current_parent_field = None
    section_had_subheader = False

    for raw_row in rows:
        row = normalize_row(raw_row)
        if not any(c is not None for c in row):
            continue

        # ── Section header detection ──────────────────────────────────────────
        sec_info = get_section_info(row, tab_keys)
        if sec_info:
            close_section(sections, current_section)
            section_name, section_type, tab_key = sec_info
            current_section = {
                "section_name": section_name,
                "section_type": section_type,
                "tab_key": tab_key,
                "entries": [],
            }
            current_section_type = section_type
            current_parent_field = infer_parent_field(section_name)
            section_had_subheader = False
            continue

        # ── Sub-header rows (skip, but note their presence) ───────────────────
        if is_column_header_row(row):
            if current_section is not None:
                section_had_subheader = True
            else:
                # Flat format: auto-open a default section
                tab_key = tab_keys[0] if len(tab_keys) == 1 else None
                current_section = {
                    "section_name": "Field Definitions",
                    "section_type": "flat",
                    "tab_key": tab_key,
                    "entries": [],
                }
                current_section_type = "flat"
                section_had_subheader = True
            continue

        if current_section is None:
            continue

        col_a = cell_str(row, 0)
        col_b = cell_str(row, 1)
        col_c = cell_str(row, 2)
        col_d = cell_str(row, 3)

        # ── Parse by section type ─────────────────────────────────────────────

        if current_section_type == "column_key_tabbed":
            # col_a = tab name hint, col_b = field name, col_c = definition
            if not col_b:
                continue
            tab_hint = slugify(col_a) if col_a else None
            row_tab = next(
                (k for k in tab_keys if k == tab_hint or (tab_hint and tab_hint in k)),
                current_section.get("tab_key"),
            )
            # Split into a new section when the tab changes
            last_tab = current_section.get("_last_tab")
            if row_tab and row_tab != last_tab:
                if current_section["entries"]:
                    close_section(sections, current_section)
                    current_section = {
                        "section_name": f"Column Key ({col_a})",
                        "section_type": "column_key_tabbed",
                        "tab_key": row_tab,
                        "entries": [],
                    }
                else:
                    current_section["tab_key"] = row_tab
                current_section["_last_tab"] = row_tab

            entry = {
                "name": col_b,
                "definition": col_c or None,
                "notes": None,
                "entry_type": "field",
                "parent_field": None,
            }

        elif current_section_type == "columns":
            # col_a = field name, col_b = units, col_c = definition, col_d = notes
            if not col_a or not col_c:
                continue
            units_note = f"Units: {col_b}." if col_b and col_b.lower() not in ("n/a", "") else ""
            notes = f"{units_note} {col_d}".strip() or None
            if not current_section["tab_key"]:
                norm = normalize(col_a)
                if norm in all_col_norms:
                    current_section["tab_key"] = all_col_norms[norm]
            entry = {
                "name": col_a,
                "definition": col_c,
                "notes": notes,
                "entry_type": "field",
                "parent_field": None,
            }

        elif current_section_type == "definitions":
            if section_had_subheader:
                # 3-col: col_a = parent field, col_b = value name, col_c = definition
                if not col_b:
                    continue
                if col_a:
                    current_parent_field = col_a
                entry = {
                    "name": col_b,
                    "definition": col_c or None,
                    "notes": col_d or None,
                    "entry_type": "definition",
                    "parent_field": current_parent_field or col_a,
                }
            else:
                # 2-col: col_a = value name, col_b = definition
                if not col_a or not col_b:
                    continue
                entry = {
                    "name": col_a,
                    "definition": col_b,
                    "notes": col_c or None,
                    "entry_type": "definition",
                    "parent_field": current_parent_field,
                }

        elif current_section_type == "flat":
            # col_a = field name, col_b = definition
            if not col_a or not col_b:
                continue
            entry = {
                "name": col_a,
                "definition": col_b,
                "notes": col_c or None,
                "entry_type": "field",
                "parent_field": None,
            }

        else:
            # "dictionary" — original coal mine / coal plant format
            if not col_a or not col_b:
                continue
            norm_a = normalize(col_a)
            tab_key = current_section["tab_key"]
            tab_col_set = tab_col_sets.get(tab_key, set()) if tab_key else set()

            if norm_a in tab_col_set or norm_a in all_col_norms:
                entry_type = "field"
                current_parent_field = col_a
                if not tab_key and norm_a in all_col_norms:
                    current_section["tab_key"] = all_col_norms[norm_a]
            else:
                entry_type = "definition"

            entry = {
                "name": col_a,
                "definition": col_b,
                "notes": col_c or None,
                "entry_type": entry_type,
                "parent_field": current_parent_field if entry_type == "definition" else None,
            }

        current_section["entries"].append(entry)

    close_section(sections, current_section)
    return header_info, sections


# ── File processing ───────────────────────────────────────────────────────────

def process_file(filepath):
    print(f"\nProcessing: {filepath.name}")
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_names = wb.sheetnames

    metadata_tabs = [n for n in sheet_names if n.lower() in METADATA_TAB_NAMES]
    data_tabs = [n for n in sheet_names if n.lower() not in METADATA_TAB_NAMES]

    if not metadata_tabs:
        print("  No metadata tab found, skipping")
        wb.close()
        return None

    print(f"  Metadata tab(s): {metadata_tabs}")
    print(f"  Data tab(s): {data_tabs[:8]}{'...' if len(data_tabs) > 8 else ''}")

    # Build normalized column sets per data tab
    tab_col_sets = {}
    for tab_name in data_tabs:
        ws = wb[tab_name]
        for row in ws.iter_rows(values_only=True):
            cols = {
                normalize(re.sub(r"\s+", " ", str(c).strip()))
                for c in row
                if c is not None
            }
            cols.discard("")
            if cols:
                tab_col_sets[slugify(tab_name)] = cols
            break

    # Header info from the first "about/readme/notes" tab
    header_info = {"title": None, "citation": None, "contact": None, "license": None}
    header_tabs = [n for n in metadata_tabs if n.lower() in HEADER_INFO_TAB_NAMES]
    if header_tabs:
        rows = list(wb[header_tabs[0]].iter_rows(values_only=True))
        header_info = extract_header_info(rows)

    # Field definitions from ALL metadata tabs
    all_sections = []
    for tab_name in metadata_tabs:
        _, sections = parse_metadata_tab(wb[tab_name], tab_col_sets)
        all_sections.extend(sections)

    wb.close()

    n_fields = sum(
        len([e for e in s["entries"] if e["entry_type"] == "field"])
        for s in all_sections
    )
    n_defs = sum(
        len([e for e in s["entries"] if e["entry_type"] == "definition"])
        for s in all_sections
    )
    print(f"  {len(all_sections)} section(s) | {n_fields} field entries | {n_defs} definition entries")

    return {
        "source_file": filepath.name,
        "tracker_slug": slugify(filepath.stem),
        "header_info": header_info,
        "data_tabs": {slugify(t): t for t in data_tabs},
        "metadata_tabs": {slugify(t): t for t in metadata_tabs},
        "sections": all_sections,
    }


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    xlsx_files = sorted(SPREADSHEET_DIR.glob("*.xlsx"))

    if not xlsx_files:
        print(f"No .xlsx files found in {SPREADSHEET_DIR}/")
        return

    for filepath in xlsx_files:
        result = process_file(filepath)
        if result is None:
            continue

        out = OUTPUT_DIR / f"{result['tracker_slug']}.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        print(f"  → {out}")

    print("\nDone.")


if __name__ == "__main__":
    main()
